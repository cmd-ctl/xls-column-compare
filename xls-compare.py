import openpyxl

input_file = "file1.xlsx"  # input Excel
output_file = "file2.xlsx"  # save filename
column_to_check = 'G'  # check column
sheet_name = "List1" # sheet name
text_to_write = "compare confirm"  # text mark

codes_to_match = [
    "DB46", "B3DB", "A312", "34EC", "4490", "5439", "2BC0",
    "192B", "D19A", "3858", "9EB6", "CD4E", "99E8", "ADE6"
]  # list to check
result_column = 'H'  # column for mark
wb = openpyxl.load_workbook(input_file)

# check if list exists
if sheet_name not in wb.sheetnames:
    print(f"List '{sheet_name}' not found in file {input_file}.")
    exit()

sheet = wb[sheet_name]

for row in range(2, sheet.max_row + 1):  # start from second row
    cell_value = sheet[f"{column_to_check}{row}"].value
    if cell_value in codes_to_match:  # if compared
        sheet[f"{result_column}{row}"] = text_to_write  # making mark

wb.save(output_file)
print(f"File saved as {output_file}")